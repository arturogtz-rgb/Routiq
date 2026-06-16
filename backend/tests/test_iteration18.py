"""Iteration 18 backend tests:
- Catalog export (.xlsx, admin-only)
- WhatsApp <-> Quotation link by folio (link / by-quotation / chats include quotation / unlink)
- Per-tenant Gmail OAuth (authorize URL, callback bad state, disconnect, admin-only)
"""
import io
import os
import time
import uuid
import pytest
import requests
import openpyxl
from pymongo import MongoClient
from urllib.parse import urlparse, parse_qs

BASE = os.environ.get("REACT_APP_BACKEND_URL", "https://master-panel-6.preview.emergentagent.com").rstrip("/")
API = f"{BASE}/api"

ADMIN_EMAIL = "admin@aventurate.mx"
ADMIN_PASSWORD = "Demo2026!"
EXEC_EMAIL = "ejecutivo@aventurate.mx"
EXEC_PASSWORD = "Demo2026!"
BAILEYS_SECRET = "586784a4d615b6b2df4402ff599e063ef0438f3efc235d5b"

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "routiq")

UNIQUE = uuid.uuid4().hex[:8]


def _login(email, password):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=20)
    assert r.status_code == 200, f"login failed for {email}: {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="module")
def admin():
    return _login(ADMIN_EMAIL, ADMIN_PASSWORD)


@pytest.fixture(scope="module")
def executive():
    return _login(EXEC_EMAIL, EXEC_PASSWORD)


@pytest.fixture(scope="module")
def tenant_id(admin):
    r = admin.get(f"{API}/companies/me", timeout=15)
    assert r.status_code == 200
    return r.json()["id"]


@pytest.fixture(scope="module")
def mongo():
    cli = MongoClient(MONGO_URL)
    yield cli[DB_NAME]
    cli.close()


# ---------------------------------------------------------------------------
# 1) CATALOG EXPORT
# ---------------------------------------------------------------------------
class TestCatalogExport:
    def test_export_admin_returns_xlsx_with_three_sheets(self, admin):
        r = admin.get(f"{API}/catalog/export", timeout=30)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")
        assert "attachment" in r.headers.get("content-disposition", "").lower()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        for name in ("Paquetes", "Tours", "Traslados"):
            assert name in wb.sheetnames, f"missing sheet {name}, got {wb.sheetnames}"
        # Header row must match the template columns
        pkg_headers = [c.value for c in wb["Paquetes"][1]]
        assert pkg_headers[:3] == ["code", "name", "nights"]
        svc_headers = [c.value for c in wb["Tours"][1]]
        assert svc_headers == ["name", "description", "net_price", "public_price", "unit"]

    def test_export_has_tenant_data(self, admin):
        # tenant has seeded packages/services from previous iterations
        pkgs = admin.get(f"{API}/packages", timeout=15).json()
        r = admin.get(f"{API}/catalog/export", timeout=30)
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        # Paquetes sheet should have header + N rows
        rows = list(wb["Paquetes"].iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if r and r[0]]
        assert len(non_empty) == len(pkgs), f"export rows={len(non_empty)} vs api packages={len(pkgs)}"

    def test_export_executive_forbidden(self, executive):
        r = executive.get(f"{API}/catalog/export", timeout=15)
        assert r.status_code == 403, r.text


# ---------------------------------------------------------------------------
# 2) WHATSAPP <-> QUOTATION LINKING
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module")
def wa_number(admin):
    r = admin.get(f"{API}/whatsapp/numbers", timeout=15)
    assert r.status_code == 200
    nums = r.json()
    assert nums, "no whatsapp numbers seeded for tenant"
    # Prefer 'Ventas GDL' seed if present
    seed = next((n for n in nums if "Ventas" in (n.get("label") or "")), nums[0])
    return seed


@pytest.fixture(scope="module")
def quotation(admin):
    r = admin.get(f"{API}/quotations", timeout=15)
    assert r.status_code == 200, r.text
    quos = r.json()
    assert quos, "no quotations available for tenant — seed needed"
    return quos[0]


@pytest.fixture(scope="module")
def injected_chat(tenant_id, wa_number, mongo):
    """Inject a chat via webhook so /api/whatsapp/chats has data."""
    chat_id = f"5213311{UNIQUE[:7]}@s.whatsapp.net"
    msg_id = f"TEST_ITER18_{UNIQUE}_{int(time.time())}"
    payload = {
        "event": "message",
        "session_id": f"{tenant_id}_{wa_number['id']}",
        "chat_id": chat_id,
        "message_id": msg_id,
        "from_me": False,
        "text": "Hola, me interesa la cotización",
        "push_name": "Test Iter18",
        "timestamp": int(time.time()),
    }
    r = requests.post(f"{API}/whatsapp/webhook", json=payload,
                      headers={"x-baileys-secret": BAILEYS_SECRET}, timeout=15)
    assert r.status_code == 200, r.text
    yield {"chat_id": chat_id, "message_id": msg_id}
    # Teardown - delete messages and any link
    mongo.whatsapp_messages.delete_many({"message_id": msg_id})
    mongo.whatsapp_messages.delete_many({"chat_id": chat_id, "tenant_id": tenant_id})


class TestWhatsAppLink:
    def test_webhook_unauthorized_without_secret(self):
        r = requests.post(f"{API}/whatsapp/webhook", json={"event": "ping"}, timeout=10)
        assert r.status_code == 401

    def test_chats_list_includes_injected(self, admin, wa_number, injected_chat):
        r = admin.get(f"{API}/whatsapp/chats?number_id={wa_number['id']}", timeout=15)
        assert r.status_code == 200, r.text
        chats = r.json()
        match = [c for c in chats if c["chat_id"] == injected_chat["chat_id"]]
        assert match, f"injected chat not found in {len(chats)} chats"
        # Not linked yet
        assert match[0]["quotation_id"] is None
        assert match[0]["quotation_code"] is None

    def test_link_quotation_not_found(self, admin, wa_number, injected_chat):
        r = admin.post(f"{API}/whatsapp/link", json={
            "quotation_id": "nonexistent-quotation-id-xxx",
            "number_id": wa_number["id"],
            "chat_id": injected_chat["chat_id"],
        }, timeout=15)
        assert r.status_code == 404, r.text

    def test_link_create_and_by_quotation(self, admin, wa_number, quotation, injected_chat, mongo, tenant_id):
        # link
        r = admin.post(f"{API}/whatsapp/link", json={
            "quotation_id": quotation["id"],
            "number_id": wa_number["id"],
            "chat_id": injected_chat["chat_id"],
        }, timeout=15)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["ok"] is True
        assert data["quotation_code"] == quotation.get("code", "")
        assert data["phone"] == injected_chat["chat_id"].split("@")[0]

        # by-quotation
        r2 = admin.get(f"{API}/whatsapp/links/by-quotation/{quotation['id']}", timeout=15)
        assert r2.status_code == 200, r2.text
        link = r2.json()
        assert link.get("phone") == injected_chat["chat_id"].split("@")[0]
        assert link.get("quotation_code") == quotation.get("code", "")
        assert link.get("chat_id") == injected_chat["chat_id"]

        # chats now expose quotation_id/code
        r3 = admin.get(f"{API}/whatsapp/chats?number_id={wa_number['id']}", timeout=15)
        chats = r3.json()
        match = [c for c in chats if c["chat_id"] == injected_chat["chat_id"]]
        assert match
        assert match[0]["quotation_id"] == quotation["id"]
        assert match[0]["quotation_code"] == quotation.get("code", "")

    def test_link_replace_on_same_quotation(self, admin, wa_number, quotation, injected_chat, tenant_id, mongo):
        # Posting again with same quotation+chat must not create duplicates
        admin.post(f"{API}/whatsapp/link", json={
            "quotation_id": quotation["id"],
            "number_id": wa_number["id"],
            "chat_id": injected_chat["chat_id"],
        }, timeout=15)
        count = mongo.whatsapp_links.count_documents({
            "tenant_id": tenant_id, "quotation_id": quotation["id"],
        })
        assert count == 1, f"expected 1 link, got {count}"

    def test_unlink(self, admin, quotation):
        r = admin.delete(f"{API}/whatsapp/link/{quotation['id']}", timeout=15)
        assert r.status_code == 200, r.text
        assert r.json()["ok"] is True
        # by-quotation returns empty
        r2 = admin.get(f"{API}/whatsapp/links/by-quotation/{quotation['id']}", timeout=15)
        assert r2.status_code == 200
        assert r2.json() == {}


# ---------------------------------------------------------------------------
# 3) GMAIL OAUTH (per-tenant)
# ---------------------------------------------------------------------------
class TestGmailOAuth:
    def test_authorize_without_creds_returns_400(self, admin, mongo, tenant_id):
        # ensure tenant has no gmail object first
        mongo.companies.update_one({"id": tenant_id}, {"$unset": {"gmail": ""}})
        r = admin.get(f"{API}/oauth/gmail/authorize", timeout=15)
        assert r.status_code == 400, r.text

    def test_patch_integrations_persists_gmail_creds(self, admin):
        payload = {
            "email_provider": "gmail",
            "gmail_client_id": f"TEST_ITER18_{UNIQUE}.apps.googleusercontent.com",
            "gmail_client_secret": f"GOCSPX-test-{UNIQUE}",
            "gmail_from_name": "Aventúrate Test",
        }
        r = admin.patch(f"{API}/companies/me/integrations", json=payload, timeout=15)
        assert r.status_code == 200, r.text
        view = r.json()
        assert view.get("gmail_client_id_set") is True
        assert view.get("gmail_client_secret_set") is True
        assert view.get("gmail_connected") is False
        assert view.get("email_provider") == "gmail"

    def test_authorize_with_creds_returns_google_url(self, admin):
        r = admin.get(f"{API}/oauth/gmail/authorize", timeout=15)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "url" in data and "redirect_uri" in data
        url = data["url"]
        assert url.startswith("https://accounts.google.com/o/oauth2/v2/auth")
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        assert "gmail.send" in qs.get("scope", [""])[0]
        assert qs.get("access_type", [""])[0] == "offline"
        assert qs.get("prompt", [""])[0] == "consent"
        assert qs.get("state", [""])[0]
        assert qs.get("response_type", [""])[0] == "code"
        assert data["redirect_uri"].endswith("/api/oauth/gmail/callback")
        assert "routiq-planes.preview.emergentagent.com" in data["redirect_uri"]

    def test_authorize_executive_forbidden(self, executive):
        r = executive.get(f"{API}/oauth/gmail/authorize", timeout=15)
        assert r.status_code == 403

    def test_callback_invalid_state_redirects_to_error(self, admin):
        # follow_redirects=False to see the 307
        r = admin.get(f"{API}/oauth/gmail/callback?state=BADSTATE&code=x",
                      allow_redirects=False, timeout=15)
        assert r.status_code in (302, 307), r.text
        loc = r.headers.get("location", "")
        assert "/app/settings" in loc and "gmail=error" in loc
        assert r.status_code != 500

    def test_disconnect_admin(self, admin):
        r = admin.post(f"{API}/oauth/gmail/disconnect", timeout=15)
        assert r.status_code == 200, r.text
        view = r.json()
        assert view.get("gmail_connected") is False

    def test_disconnect_executive_forbidden(self, executive):
        r = executive.post(f"{API}/oauth/gmail/disconnect", timeout=15)
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# CLEANUP (module teardown via finalizer)
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module", autouse=True)
def cleanup_demo(request, mongo, tenant_id):
    yield
    # Restore tenant: email_provider=resend, no gmail object, no test whatsapp data
    mongo.companies.update_one(
        {"id": tenant_id},
        {"$set": {"email_provider": "resend"}, "$unset": {"gmail": ""}},
    )
    mongo.whatsapp_messages.delete_many({"message_id": {"$regex": f"^TEST_ITER18_{UNIQUE}"}})
    mongo.whatsapp_messages.delete_many({"chat_id": {"$regex": f"^5213311{UNIQUE[:7]}"}})
    mongo.whatsapp_links.delete_many({"chat_id": {"$regex": f"^5213311{UNIQUE[:7]}"}})
