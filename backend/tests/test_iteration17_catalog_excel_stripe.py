"""Iteration 17 — Bulk catalog Excel import + Stripe secret clear.

Covers:
 - GET /api/catalog/template returns .xlsx (admin-only, 4 sheets)
 - POST /api/catalog/import validates row-by-row (valid create / invalid report)
 - Non .xlsx file → 400
 - executive role → 403 on template/import/stripe-secret-delete
 - PATCH /api/companies/me/integrations sets stripe_secret_set=true
 - DELETE /api/companies/me/integrations/stripe-secret → stripe_secret_set=false, stripe_enabled=false
 - Multi-tenant isolation: imported docs carry the admin's tenant_id

Cleanup: deletes anything with code/name prefix "TEST_ITER17_" and resets
stripe to no secret + disabled.
"""
import io
import os
import uuid

import openpyxl
import pytest
import requests

def _load_base_url():
    url = os.environ.get("REACT_APP_BACKEND_URL", "")
    if not url:
        try:
            with open("/app/frontend/.env") as f:
                for line in f:
                    if line.strip().startswith("REACT_APP_BACKEND_URL="):
                        url = line.split("=", 1)[1].strip().strip('"').strip("'")
                        break
        except Exception:
            pass
    return url.rstrip("/")


BASE_URL = _load_base_url()
API = f"{BASE_URL}/api"

ADMIN = {"email": "admin@aventurate.mx", "password": "Demo2026!"}
EXEC = {"email": "ejecutivo@aventurate.mx", "password": "Demo2026!"}

PREFIX = "TEST_ITER17_"
PKG_COLS = ["code", "name", "nights", "description", "includes", "excludes",
            "hotel_name", "sencilla", "doble", "triple", "cuadruple", "menor"]
SVC_COLS = ["name", "description", "net_price", "public_price", "unit"]


def _login(session, creds):
    r = session.post(f"{API}/auth/login", json=creds, timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return r.json()


@pytest.fixture(scope="module")
def admin_session():
    s = requests.Session()
    _login(s, ADMIN)
    yield s
    # cleanup: clear stripe secret + delete TEST_ITER17_ packages/services
    try:
        s.delete(f"{API}/companies/me/integrations/stripe-secret", timeout=20)
    except Exception:
        pass
    try:
        pkgs = s.get(f"{API}/packages", timeout=20).json()
        for p in pkgs:
            if str(p.get("code", "")).startswith(PREFIX) or str(p.get("name", "")).startswith(PREFIX):
                s.delete(f"{API}/packages/{p['id']}", timeout=20)
    except Exception:
        pass
    try:
        svcs = s.get(f"{API}/services", timeout=20).json()
        for sv in svcs:
            if str(sv.get("name", "")).startswith(PREFIX):
                s.delete(f"{API}/services/{sv['id']}", timeout=20)
    except Exception:
        pass


@pytest.fixture(scope="module")
def exec_session():
    s = requests.Session()
    _login(s, EXEC)
    yield s


@pytest.fixture(scope="module")
def tenant_id(admin_session):
    r = admin_session.get(f"{API}/companies/me", timeout=20)
    assert r.status_code == 200
    return r.json()["id"]


# ---------------------------------------------------------------------------
# Template download
# ---------------------------------------------------------------------------
class TestTemplate:
    def test_admin_downloads_xlsx_template_4_sheets(self, admin_session):
        r = admin_session.get(f"{API}/catalog/template", timeout=20)
        assert r.status_code == 200, r.text
        ctype = r.headers.get("content-type", "")
        assert "spreadsheetml" in ctype, f"unexpected content-type: {ctype}"
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        names = set(wb.sheetnames)
        for needed in ("Paquetes", "Tours", "Traslados", "Instrucciones"):
            assert needed in names, f"missing sheet {needed} in {names}"

    def test_executive_template_403(self, exec_session):
        r = exec_session.get(f"{API}/catalog/template", timeout=20)
        assert r.status_code == 403, f"expected 403 got {r.status_code}"


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
def _build_xlsx(valid_code, valid_name, dup_code):
    """Builds a workbook with mixed valid + invalid rows."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Paquetes"
    ws1.append(PKG_COLS)
    # row 2 — valid
    ws1.append([valid_code, valid_name, 3, "Desc", "Hospedaje;Desayuno", "Vuelos",
                "Hotel Test", 1000, 800, 700, 600, 400])
    # row 3 — invalid: missing name
    ws1.append([f"{PREFIX}PKG-MISS", "", 2, "", "", "", "", 0, 0, 0, 0, 0])
    # row 4 — invalid: nights=0
    ws1.append([f"{PREFIX}PKG-BADN", f"{PREFIX}Bad Nights", 0, "", "", "", "", 0, 0, 0, 0, 0])
    # row 5 — invalid: duplicate code (will collide with row 2 OR with existing pkg)
    ws1.append([valid_code, f"{PREFIX}Dup Code", 1, "", "", "", "", 0, 0, 0, 0, 0])

    ws2 = wb.create_sheet("Tours")
    ws2.append(SVC_COLS)
    # row 2 — valid tour with public_price=0 (autocalc expected)
    ws2.append([f"{PREFIX}Tour OK", "Recorrido", 600, 0, "per_person"])
    # row 3 — invalid: missing name
    ws2.append(["", "no name", 100, 0, "per_group"])
    # row 4 — valid tour with bad unit → defaults to per_group
    ws2.append([f"{PREFIX}Tour BadUnit", "x", 200, 300, "weird_unit"])

    ws3 = wb.create_sheet("Traslados")
    ws3.append(SVC_COLS)
    # row 2 — valid traslado
    ws3.append([f"{PREFIX}Traslado OK", "Sedán", 500, 0, "per_group"])
    # row 3 — invalid: net_price not numeric
    ws3.append([f"{PREFIX}Traslado Bad", "x", "abc", 0, "per_group"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestImport:
    def test_non_xlsx_returns_400(self, admin_session):
        files = {"file": ("notes.txt", b"hola", "text/plain")}
        r = admin_session.post(f"{API}/catalog/import", files=files, timeout=30)
        assert r.status_code == 400, f"got {r.status_code}: {r.text}"

    def test_executive_import_403(self, exec_session):
        wb = openpyxl.Workbook()
        wb.active.append(PKG_COLS)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("e.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = exec_session.post(f"{API}/catalog/import", files=files, timeout=30)
        assert r.status_code == 403, f"expected 403 got {r.status_code}"

    def test_admin_import_mixed_valid_and_invalid(self, admin_session, tenant_id):
        suffix = uuid.uuid4().hex[:6].upper()
        valid_code = f"{PREFIX}PKG-{suffix}"
        valid_name = f"{PREFIX}Pkg Valid {suffix}"
        buf = _build_xlsx(valid_code, valid_name, valid_code)
        files = {"file": ("test.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_session.post(f"{API}/catalog/import", files=files, timeout=60)
        assert r.status_code == 200, f"got {r.status_code}: {r.text}"
        data = r.json()
        # structure
        assert data["ok"] is True
        assert set(data["imported"].keys()) == {"paquetes", "tours", "traslados"}
        # expected: 1 valid pkg, 2 valid tours, 1 valid traslado → 4 imported
        assert data["imported"]["paquetes"] == 1, data
        assert data["imported"]["tours"] == 2, data
        assert data["imported"]["traslados"] == 1, data
        assert data["total_imported"] == 4
        # expected errors: 3 in Paquetes (rows 3,4,5) + 1 in Tours (row 3) + 1 in Traslados (row 3)
        assert data["error_count"] == 5, data["errors"]
        rows_by_sheet = {(e["sheet"], e["row"]) for e in data["errors"]}
        for needed in {("Paquetes", 3), ("Paquetes", 4), ("Paquetes", 5),
                       ("Tours", 3), ("Traslados", 3)}:
            assert needed in rows_by_sheet, f"missing error row {needed} in {rows_by_sheet}"
        # each error must contain a message
        for e in data["errors"]:
            assert e.get("message"), e

        # GET /packages should contain valid_code
        pkgs = admin_session.get(f"{API}/packages", timeout=20).json()
        match = [p for p in pkgs if p.get("code") == valid_code]
        assert len(match) == 1, f"valid pkg not found in /packages"
        assert match[0]["tenant_id"] == tenant_id  # tenant isolation
        assert match[0]["name"] == valid_name

        # GET /services should contain the imported tour with autocalc public_price
        svcs = admin_session.get(f"{API}/services", timeout=20).json()
        tour_ok = [s for s in svcs if s.get("name") == f"{PREFIX}Tour OK"]
        assert len(tour_ok) == 1
        assert tour_ok[0]["tenant_id"] == tenant_id
        assert tour_ok[0]["category"] == "tour"
        assert tour_ok[0]["net_price"] == 600
        # public_price should be > 0 (autocalculated when input was 0)
        assert tour_ok[0]["public_price"] > 0, tour_ok[0]
        assert tour_ok[0]["unit"] == "per_person"

        # bad unit → defaulted to per_group
        bad_unit = [s for s in svcs if s.get("name") == f"{PREFIX}Tour BadUnit"]
        assert len(bad_unit) == 1
        assert bad_unit[0]["unit"] == "per_group"

        # traslado valid present and categorized correctly
        traslado = [s for s in svcs if s.get("name") == f"{PREFIX}Traslado OK"]
        assert len(traslado) == 1
        assert traslado[0]["category"] == "traslado"


# ---------------------------------------------------------------------------
# Stripe clear
# ---------------------------------------------------------------------------
class TestStripeClear:
    def test_set_and_clear_stripe_secret(self, admin_session):
        # set a fake secret
        r = admin_session.patch(
            f"{API}/companies/me/integrations",
            json={"stripe_secret_key": "sk_test_ITER17_FAKE"},
            timeout=20,
        )
        assert r.status_code == 200, r.text
        assert r.json().get("stripe_secret_set") is True

        # clear it
        r2 = admin_session.delete(f"{API}/companies/me/integrations/stripe-secret", timeout=20)
        assert r2.status_code == 200, r2.text
        body = r2.json()
        assert body.get("stripe_secret_set") is False
        assert body.get("stripe_enabled") is False

        # confirm via GET
        r3 = admin_session.get(f"{API}/companies/me/integrations", timeout=20)
        assert r3.status_code == 200
        d = r3.json()
        assert d.get("stripe_secret_set") is False
        assert d.get("stripe_enabled") is False

    def test_executive_cannot_clear_stripe_secret(self, exec_session):
        r = exec_session.delete(f"{API}/companies/me/integrations/stripe-secret", timeout=20)
        assert r.status_code == 403, f"expected 403, got {r.status_code}: {r.text}"
