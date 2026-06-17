"""Iteration 32 API regression: sales report config, manual trigger, multi-currency public quotation, email attachment plumbing."""
import os
import io
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = {"email": "admin@aventurate.mx", "password": "Demo2026!"}
EXEC = {"email": "ejecutivo@aventurate.mx", "password": "Demo2026!"}


def _login(creds):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json=creds, timeout=20)
    assert r.status_code == 200, f"login failed {r.status_code}: {r.text}"
    return s


@pytest.fixture(scope="module")
def admin_session():
    return _login(ADMIN)


@pytest.fixture(scope="module")
def exec_session():
    return _login(EXEC)


# ---------- Sales report config persistence ----------
class TestSalesReportConfig:
    def test_get_returns_report_fields(self, admin_session):
        r = admin_session.get(f"{API}/companies/me/integrations", timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        for f in ("report_enabled", "report_frequency", "report_day", "report_hour"):
            assert f in data, f"missing field {f} in {list(data.keys())}"

    def test_patch_weekly_persists(self, admin_session):
        payload = {"report_enabled": True, "report_frequency": "weekly", "report_day": 1, "report_hour": 8}
        r = admin_session.patch(f"{API}/companies/me/integrations", json=payload, timeout=20)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["report_enabled"] is True
        assert d["report_frequency"] == "weekly"
        assert d["report_day"] == 1
        assert d["report_hour"] == 8
        # GET reflects
        r2 = admin_session.get(f"{API}/companies/me/integrations", timeout=20)
        d2 = r2.json()
        assert d2["report_enabled"] is True
        assert d2["report_frequency"] == "weekly"
        assert d2["report_day"] == 1
        assert d2["report_hour"] == 8

    def test_patch_monthly_persists(self, admin_session):
        payload = {"report_frequency": "monthly", "report_day": 15, "report_hour": 9}
        r = admin_session.patch(f"{API}/companies/me/integrations", json=payload, timeout=20)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["report_frequency"] == "monthly"
        assert d["report_day"] == 15
        assert d["report_hour"] == 9
        r2 = admin_session.get(f"{API}/companies/me/integrations", timeout=20)
        d2 = r2.json()
        assert d2["report_frequency"] == "monthly"
        assert d2["report_day"] == 15
        assert d2["report_hour"] == 9


# ---------- Manual report trigger ----------
class TestManualReportTrigger:
    def test_admin_send_report_graceful(self, admin_session):
        r = admin_session.post(f"{API}/stats/sales/send-report?period=week", timeout=60)
        # In this preview no email provider key configured -> expected ok:false JSON, HTTP 200
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"
        body = r.json()
        assert "ok" in body
        # Expect false because no provider configured
        if body.get("ok") is False:
            assert "detail" in body and isinstance(body["detail"], str) and len(body["detail"]) > 0
            # Spanish message
            assert any(kw in body["detail"].lower() for kw in ["correo", "proveedor", "no se pudo", "rechaz", "smtp", "resend"])
        else:
            # If provider somehow worked, ensure expected keys
            assert "to" in body

    def test_executive_forbidden(self, exec_session):
        r = exec_session.post(f"{API}/stats/sales/send-report?period=week", timeout=20)
        assert r.status_code == 403, f"Expected 403, got {r.status_code}: {r.text}"


# ---------- Multi-currency public quotation ----------
class TestMultiCurrencyPublic:
    @pytest.fixture(scope="class")
    def base_state(self, admin_session):
        """Snapshot the original base_currency to restore at end of class."""
        r = admin_session.get(f"{API}/companies/me/integrations", timeout=20)
        original = r.json().get("base_currency") or "MXN"
        yield original
        # Always restore to MXN as required by spec
        admin_session.patch(f"{API}/companies/me/integrations", json={"base_currency": "MXN"}, timeout=20)

    def test_public_quotation_multicurrency_flip(self, admin_session, base_state):
        # 1) Ensure base is MXN
        admin_session.patch(f"{API}/companies/me/integrations", json={"base_currency": "MXN"}, timeout=20)

        # 2) Get a quotation id
        r = admin_session.get(f"{API}/quotations", timeout=20)
        assert r.status_code == 200, r.text
        items = r.json()
        # endpoint may return list directly or {items: [...]}
        if isinstance(items, dict):
            items = items.get("items") or items.get("data") or []
        assert isinstance(items, list) and len(items) > 0, "no quotations available"
        qid = items[0]["id"]

        # 3) Create public link
        rl = admin_session.post(f"{API}/quotations/{qid}/public-link", timeout=20)
        assert rl.status_code in (200, 201), rl.text
        token = rl.json().get("token") or rl.json().get("public_token")
        assert token, f"no token in response: {rl.json()}"

        # 4) GET public quotation with MXN base -> equivalent should be USD
        rp = requests.get(f"{API}/public/quotations/{token}", timeout=20)
        assert rp.status_code == 200, rp.text
        pq = rp.json()
        pay = pq.get("payment") or {}
        for k in ("base_currency", "equivalent_amount", "equivalent_currency", "rate_mxn_per_usd"):
            assert k in pay, f"missing {k} in payment block: {list(pay.keys())}"
        assert pay["base_currency"] == "MXN", f"base_currency={pay['base_currency']}"
        assert pay["equivalent_currency"] == "USD", f"equivalent_currency={pay['equivalent_currency']}"
        assert isinstance(pay["equivalent_amount"], (int, float)) and pay["equivalent_amount"] > 0
        assert isinstance(pay["rate_mxn_per_usd"], (int, float)) and pay["rate_mxn_per_usd"] > 0
        usd_amount_first = pay["equivalent_amount"]

        # 5) Flip base_currency to USD
        rpatch = admin_session.patch(f"{API}/companies/me/integrations", json={"base_currency": "USD"}, timeout=20)
        assert rpatch.status_code == 200, rpatch.text
        assert rpatch.json().get("base_currency") == "USD"

        # 6) Re-fetch public quotation -> equivalent should now be MXN
        rp2 = requests.get(f"{API}/public/quotations/{token}", timeout=20)
        assert rp2.status_code == 200, rp2.text
        pay2 = (rp2.json().get("payment") or {})
        assert pay2["base_currency"] == "USD"
        assert pay2["equivalent_currency"] == "MXN"
        assert isinstance(pay2["equivalent_amount"], (int, float)) and pay2["equivalent_amount"] > 0
        assert pay2["rate_mxn_per_usd"] > 0

        # Sanity: USD->MXN equivalent should be larger than the MXN->USD one
        # (because 1 USD ~ 17-20 MXN); just ensure they differ
        assert pay2["equivalent_amount"] != usd_amount_first

        # 7) Restore to MXN (also done by fixture teardown for safety)
        admin_session.patch(f"{API}/companies/me/integrations", json={"base_currency": "MXN"}, timeout=20)


# ---------- Email attachment plumbing regression ----------
class TestEmailRegression:
    def test_test_resend_without_key_returns_400(self, admin_session):
        # First clear resend key by sending empty (frontend never sends •• masked)
        # Try with no payload-provided key; backend should reject if none stored either.
        # We don't want to wipe a stored key — just attempt; if a key happens to be stored,
        # this would actually try to send. To check the plumbing/Spanish message we send
        # an explicit empty key via payload — backend treats empty as missing.
        r = admin_session.post(f"{API}/companies/me/test-resend",
                               json={"resend_api_key": "", "resend_from_email": "", "resend_from_name": "", "to_email": ""},
                               timeout=20)
        # Accept 400 (no key/from) — the Spanish msg must mention key or remitente or correo
        assert r.status_code == 400, f"Expected 400, got {r.status_code}: {r.text}"
        detail = (r.json() or {}).get("detail", "")
        assert isinstance(detail, str) and len(detail) > 0
        assert any(kw in detail.lower() for kw in ["api key", "resend", "remitente", "correo", "falta"])

    def test_stats_sales_still_works(self, admin_session):
        r = admin_session.get(f"{API}/stats/sales?period=month", timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        for k in ("period", "revenue_total", "trend", "conversion", "executives", "clients", "packages", "services", "lost"):
            assert k in d, f"missing {k}"
        assert d["period"] == "month"

    def test_stats_sales_export_still_works(self, admin_session):
        r = admin_session.get(f"{API}/stats/sales/export?period=month", timeout=60)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "xlsx" in ct or "officedocument" in ct, f"unexpected content-type: {ct}"
        assert len(r.content) > 1000
        # Validate workbook opens and has the required sheets
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        for s in ["Resumen", "Ejecutivos", "Clientes", "Paquetes", "Servicios", "Perdidas"]:
            assert s in wb.sheetnames, f"missing sheet {s} in {wb.sheetnames}"
