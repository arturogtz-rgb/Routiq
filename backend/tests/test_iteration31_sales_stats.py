"""Iteration 31 — Sales & Statistics module + lost reason capture.

Covers:
- GET /api/stats/sales (week/month/quarter/year) auth + payload shape + trend length
- Role enforcement: 401 unauthenticated, 403 for executive
- GET /api/stats/sales/export returns valid xlsx with required sheets
- PATCH /api/quotations/{id}/state with reason -> persisted lost_reason -> shows in /stats/sales lost
- Regression: GET /api/catalog/analytics still returns 200 for company_admin
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://master-panel-6.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN = {"email": "admin@aventurate.mx", "password": "Demo2026!"}
EXEC = {"email": "ejecutivo@aventurate.mx", "password": "Demo2026!"}


def _login(creds):
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json=creds, timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="module")
def admin():
    return _login(ADMIN)


@pytest.fixture(scope="module")
def execu():
    return _login(EXEC)


# ---------- Sales stats ----------
EXPECTED_TREND = {"week": 7, "month": 30, "quarter": 13, "year": 12}


@pytest.mark.parametrize("period,length", list(EXPECTED_TREND.items()))
def test_sales_stats_periods(admin, period, length):
    r = admin.get(f"{API}/stats/sales", params={"period": period}, timeout=30)
    assert r.status_code == 200, r.text
    d = r.json()
    # Shape
    for k in ["period", "days", "currency", "revenue_total", "collected_total",
              "trend", "conversion", "executives", "clients", "packages", "services", "lost"]:
        assert k in d, f"missing key {k} in payload for period={period}"
    assert d["period"] == period
    assert isinstance(d["trend"], list)
    assert len(d["trend"]) == length, f"trend length for {period}: got {len(d['trend'])}, expected {length}"
    # Each trend bucket
    for b in d["trend"]:
        assert "label" in b and "revenue" in b and "count" in b
    # Conversion shape
    for k in ["total", "won", "lost", "rate"]:
        assert k in d["conversion"]


def test_sales_stats_unauth():
    r = requests.get(f"{API}/stats/sales", params={"period": "month"}, timeout=15)
    assert r.status_code == 401, f"expected 401 unauth, got {r.status_code}"


def test_sales_stats_forbidden_for_executive(execu):
    r = execu.get(f"{API}/stats/sales", params={"period": "month"}, timeout=20)
    assert r.status_code == 403, f"expected 403 for executive, got {r.status_code} {r.text}"


# ---------- Excel export ----------
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_xlsx_valid_workbook(admin):
    r = admin.get(f"{API}/stats/sales/export", params={"period": "month"}, timeout=60)
    assert r.status_code == 200, r.text
    assert r.headers.get("content-type", "").startswith(XLSX_MIME)
    cd = r.headers.get("content-disposition", "")
    assert "attachment" in cd and "routiq-ventas-month-" in cd and cd.endswith('.xlsx"'), cd
    wb = load_workbook(io.BytesIO(r.content))
    expected = {"Resumen", "Ejecutivos", "Clientes", "Paquetes", "Servicios", "Perdidas"}
    assert expected.issubset(set(wb.sheetnames)), f"missing sheets: {expected - set(wb.sheetnames)}"


def test_export_forbidden_for_executive(execu):
    r = execu.get(f"{API}/stats/sales/export", params={"period": "month"}, timeout=20)
    assert r.status_code == 403


# ---------- Lost reason capture ----------
def test_lost_reason_persistence_and_appears_in_stats(admin):
    # Pick a quotation. Prefer one already 'perdida' or any not-archived one
    r = admin.get(f"{API}/quotations", timeout=20)
    assert r.status_code == 200, r.text
    items = r.json()
    assert items, "No quotations available to test lost reason"
    q = items[0]
    qid = q["id"]
    prev_state = q.get("state")

    reason = "precio fuera de presupuesto TEST_ITER31"
    r = admin.patch(f"{API}/quotations/{qid}/state",
                    json={"state": "perdida", "reason": reason}, timeout=20)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body.get("state") == "perdida"

    # Verify it appears in stats lost list with the reason
    r = admin.get(f"{API}/stats/sales", params={"period": "month"}, timeout=30)
    assert r.status_code == 200
    lost = r.json().get("lost", [])
    code = q.get("code")
    matching = [l for l in lost if l.get("code") == code]
    assert matching, f"Quotation {code} not present in lost list"
    assert matching[0].get("reason") == reason, f"reason mismatch: {matching[0]}"

    # Best-effort restore previous state to keep dataset stable
    if prev_state and prev_state != "perdida":
        admin.patch(f"{API}/quotations/{qid}/state", json={"state": prev_state}, timeout=20)


# ---------- Regression ----------
def test_catalog_analytics_regression(admin):
    r = admin.get(f"{API}/catalog/analytics", params={"period": "month"}, timeout=20)
    assert r.status_code == 200, r.text
    d = r.json()
    # minimum shape — endpoint must keep responding
    assert isinstance(d, dict)
