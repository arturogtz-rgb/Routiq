"""Lote E — Selector de mes calendario en Ventas + Filtro por rango de fechas en cotizaciones.
Regresión rápida de precios de paquete y confirmación de reserva.
"""
import io
import os
import openpyxl
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://price-sync-alert.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@aventurate.mx"
ADMIN_PWD = "Demo2026!"


# ---------- fixtures ----------
@pytest.fixture(scope="session")
def admin_session():
    s = requests.Session()
    r = s.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PWD}, timeout=15)
    assert r.status_code == 200, r.text
    return s


# ---------- Lote E · stats?month=YYYY-MM ----------
class TestStatsMonthPicker:
    def test_stats_month_2026_06(self, admin_session):
        r = admin_session.get(f"{API}/stats/sales", params={"month": "2026-06"}, timeout=20)
        assert r.status_code == 200, r.text
        d = r.json()
        # label & month
        assert d["month"] == "2026-06"
        assert d["label"] == "Junio 2026"
        # trend buckets = número de días del mes calendario (30 para junio)
        assert isinstance(d["trend"], list)
        assert len(d["trend"]) == 30, f"esperado 30 buckets, got {len(d['trend'])}"
        # previous debe existir con estructura
        assert "previous" in d
        assert "revenue_total" in d["previous"]
        assert "conversion" in d["previous"]
        # deltas debe existir
        assert "deltas" in d
        for k in ("revenue", "collected", "created", "won", "rate"):
            assert k in d["deltas"]

    def test_stats_month_2026_02_28_buckets(self, admin_session):
        """Febrero 2026 (no bisiesto) = 28 buckets."""
        r = admin_session.get(f"{API}/stats/sales", params={"month": "2026-02"}, timeout=20)
        assert r.status_code == 200
        d = r.json()
        assert d["month"] == "2026-02"
        assert d["label"] == "Febrero 2026"
        assert len(d["trend"]) == 28

    def test_stats_month_july_31(self, admin_session):
        r = admin_session.get(f"{API}/stats/sales", params={"month": "2026-07"}, timeout=20)
        assert r.status_code == 200
        d = r.json()
        assert d["label"] == "Julio 2026"
        assert len(d["trend"]) == 31

    def test_stats_period_month_still_works(self, admin_session):
        """Sin ?month debe seguir devolviendo período de ventana móvil."""
        r = admin_session.get(f"{API}/stats/sales", params={"period": "month"}, timeout=20)
        assert r.status_code == 200
        d = r.json()
        assert d.get("month", "") == ""
        assert d["label"] == "Último mes"
        # Ventana móvil de 30 días
        assert len(d["trend"]) == 30

    def test_stats_period_week(self, admin_session):
        r = admin_session.get(f"{API}/stats/sales", params={"period": "week"}, timeout=20)
        assert r.status_code == 200
        d = r.json()
        assert d["label"] == "Última semana"
        assert len(d["trend"]) == 7


# ---------- Lote E · export xlsx por mes ----------
class TestStatsExportMonth:
    def test_export_month_xlsx(self, admin_session):
        r = admin_session.get(f"{API}/stats/sales/export", params={"month": "2026-06"}, timeout=30)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct, ct
        cd = r.headers.get("content-disposition", "")
        assert "2026-06" in cd, cd
        # Cargar el workbook para validar que el Resumen contiene el label del mes
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Resumen" in wb.sheetnames
        ws = wb["Resumen"]
        rows = list(ws.iter_rows(values_only=True))
        # Buscar "Período" y su valor
        period_row = next((row for row in rows if row and row[0] == "Período"), None)
        assert period_row is not None, rows
        assert period_row[1] == "Junio 2026", f"esperado 'Junio 2026', got {period_row[1]}"

    def test_export_period_still_works(self, admin_session):
        r = admin_session.get(f"{API}/stats/sales/export", params={"period": "month"}, timeout=30)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Resumen"]
        rows = list(ws.iter_rows(values_only=True))
        period_row = next((row for row in rows if row and row[0] == "Período"), None)
        assert period_row is not None
        assert period_row[1] == "Último mes"


# ---------- Lote E · quotations date_from/date_to ----------
class TestQuotationsDateRange:
    def test_quotations_month_range_2026_06(self, admin_session):
        r = admin_session.get(f"{API}/quotations",
                              params={"date_from": "2026-06-01", "date_to": "2026-06-30"},
                              timeout=20)
        assert r.status_code == 200, r.text
        items = r.json()
        assert isinstance(items, list)
        # Todas las cotizaciones devueltas deben tener created_at en 2026-06
        for it in items:
            ca = it.get("created_at", "")
            assert ca.startswith("2026-06"), f"cotización {it.get('code')} fuera de rango: {ca}"

    def test_quotations_month_range_2026_07(self, admin_session):
        r = admin_session.get(f"{API}/quotations",
                              params={"date_from": "2026-07-01", "date_to": "2026-07-31"},
                              timeout=20)
        assert r.status_code == 200
        items = r.json()
        for it in items:
            assert it.get("created_at", "").startswith("2026-07")

    def test_quotations_no_range_returns_all(self, admin_session):
        r_all = admin_session.get(f"{API}/quotations", timeout=20).json()
        r_jun = admin_session.get(f"{API}/quotations",
                                  params={"date_from": "2026-06-01", "date_to": "2026-06-30"},
                                  timeout=20).json()
        # El listado total debe ser >= al filtrado (no puede haber MENOS)
        assert len(r_all) >= len(r_jun)

    def test_quotations_range_excludes_out_of_month(self, admin_session):
        """Un rango de 3 días muy chico probablemente da 0 o pocos, pero ninguno fuera."""
        r = admin_session.get(f"{API}/quotations",
                              params={"date_from": "2020-01-01", "date_to": "2020-01-31"},
                              timeout=20)
        assert r.status_code == 200
        items = r.json()
        assert items == [] or all(x.get("created_at", "").startswith("2020-01") for x in items)


# ---------- Regresión Lotes C/D ----------
class TestRegression:
    def test_get_service_by_id(self, admin_session):
        """Lote C 8.1 — GET /services/{id} debe existir (bug de iter49 fix)."""
        list_r = admin_session.get(f"{API}/services", timeout=15)
        assert list_r.status_code == 200
        svcs = list_r.json()
        if not svcs:
            pytest.skip("No hay servicios seed para probar")
        sid = svcs[0]["id"]
        r = admin_session.get(f"{API}/services/{sid}", timeout=15)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["id"] == sid
        assert "name" in d

    def test_booking_confirmation_has_itinerary(self, admin_session):
        """Lote D 10.5 — Confirmación de Reserva debe tener sección itinerary."""
        # Buscar cualquier cotización ganada
        won_r = admin_session.get(f"{API}/quotations", params={"state": "ganada"}, timeout=15)
        assert won_r.status_code == 200
        won = won_r.json()
        if not won:
            pytest.skip("No hay cotizaciones ganadas")
        qid = won[0]["id"]
        r = admin_session.get(f"{API}/quotations/{qid}/booking-confirmation", timeout=15)
        assert r.status_code == 200, r.text
        d = r.json()
        # Debe tener itinerary (prefilled)
        assert "itinerary" in d, d.keys()

    def test_pricing_engine_not_broken(self, admin_session):
        """Motor de precios del Paquete Armado: subtotal + commission ~= total (independientemente del descuento)."""
        r = admin_session.get(f"{API}/quotations", timeout=15)
        assert r.status_code == 200
        qs = [x for x in r.json() if x.get("type") == "paquete"]
        if not qs:
            pytest.skip("No hay cotizaciones tipo paquete")
        # Probar al menos 3 con precios coherentes
        checked = 0
        for q in qs[:5]:
            sub = float(q.get("subtotal", 0) or 0)
            com = float(q.get("commission", 0) or 0)
            tot = float(q.get("total", 0) or 0)
            if sub == 0 and tot == 0:
                continue
            # Permitir 1 centavo de error de redondeo
            assert abs((sub + com) - tot) < 0.02, f"Cot {q.get('code')}: {sub}+{com} != {tot}"
            checked += 1
        assert checked >= 1, "Ninguna cotización paquete válida"
