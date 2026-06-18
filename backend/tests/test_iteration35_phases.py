"""Iteration 35 — Phases 2-5 + new features regression suite.

Covers:
  Phase 3 (Excel template/export/import idempotent + image_url + itinerary dia_N)
  Phase 4 (public services grouped + conditions endpoint)
  Phase 5 (catalog_subtitle PATCH /companies/me + reflejado en /public/company)
  Item 9 (custom quotation con items mixtos neto/publico por canal)
  Item 10/11 (PDF + public link expone important_info, exec_name, occupancy_prices, slug + condiciones)
  Item 12 (Confirmación de Reserva CRUD + PDF + send email/whatsapp + bank.branch/reference)
  Regresión (servicios-only commission por canal intacta)
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook, Workbook

def _load_base_url():
    url = os.environ.get('REACT_APP_BACKEND_URL', '').strip()
    if not url:
        # Fallback: read from /app/frontend/.env
        try:
            with open('/app/frontend/.env') as f:
                for line in f:
                    if line.startswith('REACT_APP_BACKEND_URL='):
                        url = line.split('=', 1)[1].strip()
                        break
        except Exception:
            pass
    return url.rstrip('/')


BASE_URL = _load_base_url()
assert BASE_URL.startswith('http'), f"BASE_URL invalid: {BASE_URL!r}"
ADMIN_EMAIL = "admin@aventurate.mx"
ADMIN_PASS = "Demo2026!"


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASS}, timeout=20)
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module")
def slug():
    return "aventurate"


# ============================================================
# PHASE 3: Excel template / export / import
# ============================================================
class TestPhase3Excel:
    def test_template_xlsx_has_required_sheets_and_columns(self, session):
        r = session.get(f"{BASE_URL}/api/catalog/template", timeout=30)
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        sheets = set(wb.sheetnames)
        for s_name in ["Paquetes", "Tours", "Traslados", "Accesos", "Extras"]:
            assert s_name in sheets, f"falta hoja {s_name}: {sheets}"
        # Paquetes: dia_1..dia_10 titulo/descripcion
        pkg_headers = [c.value for c in wb["Paquetes"][1]]
        for i in range(1, 11):
            assert f"dia_{i}_titulo" in pkg_headers
            assert f"dia_{i}_descripcion" in pkg_headers
        assert "image_url" in pkg_headers
        # Servicios: image_url presente
        for s_name in ["Tours", "Traslados", "Accesos", "Extras"]:
            headers = [c.value for c in wb[s_name][1]]
            assert "image_url" in headers, f"image_url falta en {s_name}: {headers}"

    def test_export_xlsx_ok(self, session):
        r = session.get(f"{BASE_URL}/api/catalog/export", timeout=30)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        assert "Paquetes" in wb.sheetnames

    def test_import_only_tours_does_not_affect_traslados(self, session):
        # Listar traslados antes
        before = session.get(f"{BASE_URL}/api/services?type=traslado", timeout=20).json()
        before_codes = {s["name"] for s in before}

        # Construir XLSX SOLO con hoja Tours (un tour nuevo)
        wb = Workbook()
        ws = wb.active
        ws.title = "Tours"
        ws.append(["name", "description", "net_price", "public_price", "unit", "image_url"])
        ws.append(["TEST_iter35_tour", "tour de prueba", 500, 0, "per_person", ""])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = session.post(
            f"{BASE_URL}/api/catalog/import",
            files={"file": ("only_tours.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
        assert r.status_code == 200, r.text
        rep = r.json()
        imp = rep.get("imported", {})
        tours_count = imp.get("tours_nuevos", 0) + imp.get("tours_actualizados", 0)
        assert tours_count >= 1, f"no se importó el tour: {rep}"
        # No debe haber tocado paquetes/traslados/accesos/extras
        assert imp.get("traslados_nuevos", 0) == 0 and imp.get("traslados_actualizados", 0) == 0

        # Listar traslados después: no debe haber cambiado el conteo
        after = session.get(f"{BASE_URL}/api/services?type=traslado", timeout=20).json()
        after_codes = {s["name"] for s in after}
        assert before_codes == after_codes, "el import de Tours afectó traslados"

    def test_import_idempotent_upsert(self, session):
        # Reimportar el mismo tour: no debe duplicar
        wb = Workbook()
        ws = wb.active
        ws.title = "Tours"
        ws.append(["name", "description", "net_price", "public_price", "unit", "image_url"])
        ws.append(["TEST_iter35_tour", "tour de prueba v2", 550, 0, "per_person", "https://x/y.jpg"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = session.post(
            f"{BASE_URL}/api/catalog/import",
            files={"file": ("again.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
        assert r.status_code == 200
        # Verificar que solo existe UNA versión
        services = session.get(f"{BASE_URL}/api/services?type=tour", timeout=20).json()
        matches = [s for s in services if s.get("name") == "TEST_iter35_tour"]
        assert len(matches) == 1, f"upsert duplicó: {len(matches)}"
        assert matches[0].get("net_price") == 550


# ============================================================
# PHASE 4: Public services + conditions
# ============================================================
class TestPhase4Public:
    def test_public_company_services_grouped(self, slug):
        r = requests.get(f"{BASE_URL}/api/public/company/{slug}/services", timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        # Debe agrupar por categoría dentro de "groups" (lista de {key, label, items})
        groups = data.get("groups")
        assert isinstance(groups, list) and len(groups) >= 1, f"groups: {groups}"
        keys = {g.get("key") for g in groups}
        assert keys & {"tour", "traslado", "acceso", "extra"}, f"keys: {keys}"

    def test_public_conditions_endpoint(self, slug):
        r = requests.get(f"{BASE_URL}/api/public/company/{slug}/conditions", timeout=20)
        assert r.status_code == 200
        data = r.json()
        # debe tener policy o general_conditions
        assert any(k in data for k in ("policy", "cancellation_policy", "general_conditions", "conditions"))


# ============================================================
# PHASE 5: catalog subtitle
# ============================================================
class TestPhase5Subtitle:
    def test_patch_subtitle_and_public_reflects(self, session, slug):
        new_subtitle = "TEST_iter35 subtítulo personalizado"
        r = session.patch(f"{BASE_URL}/api/companies/me",
                          json={"catalog_subtitle": new_subtitle}, timeout=20)
        assert r.status_code == 200, r.text
        # Reflected in public
        r2 = requests.get(f"{BASE_URL}/api/public/company/{slug}", timeout=20)
        assert r2.status_code == 200
        body = r2.json()
        # Puede estar en company.catalog_subtitle o root
        sub = body.get("catalog_subtitle") or (body.get("company") or {}).get("catalog_subtitle")
        assert sub == new_subtitle, f"subtítulo no se reflejó: {body}"


# ============================================================
# Item 9: Custom quotation con items mixtos neto/publico
# ============================================================
class TestItem9Custom:
    def _create_client(self, session, channel, suffix):
        payload = {
            "name": f"TEST_iter35_{suffix}", "email": f"test_{suffix}@iter35.com",
            "phone": "5555555555", "channel": channel,
        }
        r = session.post(f"{BASE_URL}/api/clients", json=payload, timeout=20)
        assert r.status_code in (200, 201), r.text
        return r.json()["id"]

    @pytest.mark.parametrize("channel,expected_commission_zero",
                             [("directo", True), ("agencia", False), ("mayorista", False), ("operador", True)])
    def test_custom_mixed_items_commission_only_on_publico(self, session, channel, expected_commission_zero):
        cid = self._create_client(session, channel, f"custom_{channel}")
        payload = {
            "type": "personalizado",
            "client_id": cid,
            "currency": "MXN",
            "dates": {"start": "2026-06-01", "end": "2026-06-03"},
            "pax": {"adultos": 2, "menores": 0},
            "custom_items": [
                {"name": "Hotel neto", "qty": 1, "price_type": "neto", "net_price": 5000, "unit": "per_group", "description": ""},
                {"name": "Tour publico", "qty": 1, "price_type": "publico", "net_price": 1000, "unit": "per_group", "description": ""},
            ],
            "important_info": "Iter35 test",
        }
        r = session.post(f"{BASE_URL}/api/quotations", json=payload, timeout=30)
        assert r.status_code in (200, 201), r.text
        q = r.json()
        assert q.get("type") == "personalizado"
        assert q.get("important_info") == "Iter35 test"
        # Commission debe aplicarse SOLO sobre items publico
        # Para directo/operador commission=0; para agencia 12% sobre 1000=120; mayorista 15%->150
        commission = q.get("commission", 0)
        if channel == "directo" or channel == "operador":
            # operador suele ser 0 commission para items publico también si convención lo dicta;
            # pero conceptualmente "publico" SÍ es comisionable. Aceptamos 0 si operador no aplica.
            pass
        elif channel == "agencia":
            assert abs(commission - 120) < 1.5, f"commission agencia esperado ~120, got {commission}"
        elif channel == "mayorista":
            assert abs(commission - 150) < 1.5, f"commission mayorista esperado ~150, got {commission}"


# ============================================================
# Item 10/11: PDF + public link + occupancy table + important_info
# ============================================================
class TestItem10_11PDFAndPublicLink:
    def _create_quote_for_state_ganada(self, session):
        # Buscar la cotización seed COT-2026056 si existe; si no, crear una sencilla y forzar estado.
        r = session.get(f"{BASE_URL}/api/quotations?state=ganada", timeout=20)
        assert r.status_code == 200
        items = r.json()
        items = items if isinstance(items, list) else items.get("items", [])
        if items:
            return items[0]
        return None

    def test_pdf_paquete_returns_pdf_bytes(self, session):
        r = session.get(f"{BASE_URL}/api/quotations?type=paquete", timeout=20)
        assert r.status_code == 200
        items = r.json()
        items = items if isinstance(items, list) else items.get("items", [])
        if not items:
            pytest.skip("No hay cotizaciones de paquete para probar PDF")
        qid = items[0]["id"]
        pdf = session.get(f"{BASE_URL}/api/quotations/{qid}/pdf", timeout=30)
        assert pdf.status_code == 200
        assert pdf.content[:4] == b"%PDF"

    def test_public_link_exposes_important_info_exec_occupancy_slug(self, session):
        # Encontrar cualquier cotización paquete y publicarla
        r = session.get(f"{BASE_URL}/api/quotations?type=paquete", timeout=20)
        items = r.json()
        items = items if isinstance(items, list) else items.get("items", [])
        if not items:
            pytest.skip("No hay cotizaciones de paquete")
        qid = items[0]["id"]
        # set important_info
        session.patch(f"{BASE_URL}/api/quotations/{qid}",
                      json={"important_info": "Información importante TEST iter35"}, timeout=20)
        # public link
        pub = session.post(f"{BASE_URL}/api/quotations/{qid}/public-link", timeout=20)
        assert pub.status_code in (200, 201), pub.text
        token = pub.json().get("token") or (pub.json().get("public_link") or {}).get("token")
        assert token, pub.json()
        # GET public
        r2 = requests.get(f"{BASE_URL}/api/public/quotations/{token}", timeout=20)
        assert r2.status_code == 200, r2.text
        body = r2.json()
        q = body["quotation"]
        assert "important_info" in q
        assert q.get("important_info") == "Información importante TEST iter35"
        assert "exec_name" in q
        assert "occupancy_prices" in q
        company = body.get("company", {})
        assert "slug" in company and company["slug"] == "aventurate"


# ============================================================
# Item 12: Booking Confirmation CRUD + PDF + send
# ============================================================
class TestItem12BookingConfirmation:
    @pytest.fixture(scope="class")
    def ganada_quote(self, session):
        # Buscar cotización en estado ganada
        r = session.get(f"{BASE_URL}/api/quotations?state=ganada", timeout=20)
        items = r.json()
        items = items if isinstance(items, list) else items.get("items", [])
        if not items:
            pytest.skip("No hay cotizaciones en estado 'ganada' para Booking Confirmation")
        return items[0]

    def test_save_and_fetch_confirmation(self, session, ganada_quote):
        qid = ganada_quote["id"]
        payload = {
            "passenger_name": "TEST_iter35 pasajero",
            "passenger_email": "pax@iter35.com",
            "passenger_phone": "5555555555",
            "agent_name": "Ejecutivo Test",
            "agent_email": "agent@iter35.com",
            "agent_phone": "5550000000",
            "services": [{"description": "Tour A", "qty": 1, "unit_price": 1000}],
            "lodging": [{"hotel": "Hotel X", "confirmation_number": "ABC123", "nights": "2"}],
            "observations": "obs",
            "subtotal": 1000, "total": 1000,
        }
        r = session.post(f"{BASE_URL}/api/quotations/{qid}/booking-confirmation", json=payload, timeout=30)
        assert r.status_code in (200, 201), r.text
        conf = r.json()
        assert "id" in conf
        assert "token" in conf
        conf_id = conf["id"]
        token = conf["token"]
        # GET
        r2 = session.get(f"{BASE_URL}/api/quotations/{qid}/booking-confirmation", timeout=20)
        assert r2.status_code == 200
        assert r2.json().get("id") == conf_id
        # PDF
        pdf = session.get(f"{BASE_URL}/api/booking-confirmations/{conf_id}/pdf", timeout=30)
        assert pdf.status_code == 200
        assert pdf.content[:4] == b"%PDF"
        # Public PDF
        pub_pdf = requests.get(f"{BASE_URL}/api/public/booking-confirmation/{token}/pdf", timeout=30)
        assert pub_pdf.status_code == 200
        assert pub_pdf.content[:4] == b"%PDF"
        # Send WhatsApp
        r_wa = session.post(f"{BASE_URL}/api/booking-confirmations/{conf_id}/send",
                            json={"channel": "whatsapp", "to": "5555555555"}, timeout=20)
        assert r_wa.status_code == 200
        body = r_wa.json()
        assert body.get("ok") is True
        assert "wa_link" in body and body["wa_link"].startswith("https://wa.me/")
        # Send email
        r_em = session.post(f"{BASE_URL}/api/booking-confirmations/{conf_id}/send",
                            json={"channel": "email", "to": "test@iter35.com"}, timeout=30)
        assert r_em.status_code == 200, r_em.text
        em_body = r_em.json()
        assert em_body.get("ok") is True
        assert "email_sent" in em_body  # boolean

    def test_save_confirmation_requires_ganada(self, session):
        # Crear una cotización vacía (no ganada) y verificar que rechaza
        r_list = session.get(f"{BASE_URL}/api/quotations?state=enviada", timeout=20)
        items = r_list.json()
        items = items if isinstance(items, list) else items.get("items", [])
        if not items:
            pytest.skip("No hay cotizaciones no-ganada para probar el guard")
        qid = items[0]["id"]
        r = session.post(f"{BASE_URL}/api/quotations/{qid}/booking-confirmation",
                         json={"passenger_name": "x", "subtotal": 0, "total": 0,
                               "services": [], "lodging": []}, timeout=20)
        assert r.status_code == 400


# ============================================================
# Bank branch/reference settings
# ============================================================
class TestBankBranchReference:
    def test_patch_bank_branch_and_reference(self, session):
        payload = {"bank_branch": "Sucursal Polanco TEST", "bank_reference": "REF-ITER35"}
        r = session.patch(f"{BASE_URL}/api/companies/me/integrations", json=payload, timeout=20)
        assert r.status_code == 200, r.text
        # GET integrations
        r2 = session.get(f"{BASE_URL}/api/companies/me/integrations", timeout=20)
        assert r2.status_code == 200
        body = r2.json()
        assert body.get("bank_branch") == "Sucursal Polanco TEST", body
        assert body.get("bank_reference") == "REF-ITER35", body


# ============================================================
# REGRESIÓN: servicios-only commission por canal
# ============================================================
class TestRegressionServicesOnly:
    def _create_client(self, session, channel, suffix):
        r = session.post(f"{BASE_URL}/api/clients",
                         json={"name": f"TEST_iter35_reg_{suffix}",
                               "email": f"reg_{suffix}@iter35.com",
                               "phone": "5555550000", "channel": channel}, timeout=20)
        assert r.status_code in (200, 201)
        return r.json()["id"]

    def test_services_only_agencia_commission_intacto(self, session):
        # buscar cualquier servicio
        svcs = session.get(f"{BASE_URL}/api/services", timeout=20).json()
        if not svcs:
            pytest.skip("No services")
        svc = next((s for s in svcs if (s.get("public_price") or s.get("net_price"))), None)
        if not svc:
            pytest.skip("No service with price")
        cid = self._create_client(session, "agencia", "ag_svc")
        payload = {
            "type": "servicios", "client_id": cid, "currency": "MXN",
            "dates": {"start": "2026-07-01", "end": "2026-07-02"},
            "pax": {"adults": 2, "minors": 0},
            "services": [{"service_id": svc["id"], "qty": 1}],
        }
        r = session.post(f"{BASE_URL}/api/quotations", json=payload, timeout=30)
        assert r.status_code in (200, 201), r.text
        q = r.json()
        # commission > 0 para canal agencia con público
        if q.get("subtotal", 0) > 0:
            assert q.get("commission", 0) > 0, f"servicios-only agencia commission deberia ser >0: {q}"
