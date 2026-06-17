"""Sales & statistics dashboard for company admins.

GET /api/stats/sales?period=week|month|quarter|year  -> aggregated metrics
GET /api/stats/sales/export?period=...               -> XLSX download
"""
import io
import logging
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from database import get_db
from auth import require_roles

log = logging.getLogger("routiq")
router = APIRouter()

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_SERVICE_CATS = {"tour", "traslado", "acceso", "extra"}


def _value(q: dict) -> float:
    v = q.get("final_total")
    if v is None:
        v = q.get("total", 0)
    return float(v or 0)


def _won_date(q: dict) -> str:
    return q.get("last_activity_at") or q.get("created_at") or ""


def _period_days(period: str) -> int:
    return {"week": 7, "month": 30, "quarter": 90, "year": 365}.get(period, 30)


def _buckets(period: str):
    """Return ordered list of {label, start, end} for the trend chart."""
    now = datetime.now(timezone.utc)
    out = []
    if period == "week":
        for i in range(6, -1, -1):
            d = (now - timedelta(days=i)).date()
            start = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
            out.append({"label": d.strftime("%d/%m"), "start": start.isoformat(),
                        "end": (start + timedelta(days=1)).isoformat()})
    elif period == "month":
        for i in range(29, -1, -1):
            d = (now - timedelta(days=i)).date()
            start = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
            out.append({"label": d.strftime("%d/%m"), "start": start.isoformat(),
                        "end": (start + timedelta(days=1)).isoformat()})
    elif period == "quarter":
        for i in range(12, -1, -1):
            end = now - timedelta(days=i * 7)
            start = end - timedelta(days=7)
            out.append({"label": end.strftime("%d/%m"), "start": start.isoformat(),
                        "end": end.isoformat()})
    else:  # year -> last 12 months
        y, m = now.year, now.month
        months = []
        for _ in range(12):
            months.append((y, m))
            m -= 1
            if m == 0:
                m = 12; y -= 1
        for (yy, mm) in reversed(months):
            start = datetime(yy, mm, 1, tzinfo=timezone.utc)
            nm, ny = (mm + 1, yy) if mm < 12 else (1, yy + 1)
            end = datetime(ny, nm, 1, tzinfo=timezone.utc)
            out.append({"label": f"{mm:02d}/{yy}", "start": start.isoformat(), "end": end.isoformat()})
    return out


async def _compute(db, tenant_id: str, period: str) -> dict:
    days = _period_days(period)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    quotations = await db.quotations.find(
        {"tenant_id": tenant_id, "deleted": {"$ne": True},
         "$or": [{"created_at": {"$gte": cutoff}}, {"last_activity_at": {"$gte": cutoff}}]},
        {"_id": 0}
    ).to_list(20000)
    users = await db.users.find({"tenant_id": tenant_id}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    uname = {u["id"]: u.get("name", "—") for u in users}

    created_in = [q for q in quotations if (q.get("created_at") or "") >= cutoff and not q.get("archived")]
    won_in = [q for q in quotations if q.get("state") == "ganada" and _won_date(q) >= cutoff]
    lost_in = [q for q in quotations if q.get("state") == "perdida" and _won_date(q) >= cutoff]

    currency = quotations[0].get("currency", "MXN") if quotations else "MXN"
    revenue_total = round(sum(_value(q) for q in won_in), 2)
    collected_total = round(sum(float(q.get("amount_paid", 0) or 0) for q in won_in), 2)

    # Trend
    buckets = _buckets(period)
    trend = []
    for b in buckets:
        rev = sum(_value(q) for q in won_in if b["start"] <= _won_date(q) < b["end"])
        cnt = sum(1 for q in won_in if b["start"] <= _won_date(q) < b["end"])
        trend.append({"label": b["label"], "revenue": round(rev, 2), "count": cnt})

    # Conversion (cohort created in period)
    total_created = len(created_in)
    won_created = sum(1 for q in created_in if q.get("state") == "ganada")
    lost_created = sum(1 for q in created_in if q.get("state") == "perdida")
    conv_rate = round(100.0 * won_created / total_created, 1) if total_created else 0.0

    # Executives
    ex = {}
    for q in created_in:
        a = q.get("assigned_to") or q.get("created_by")
        if not a:
            continue
        ex.setdefault(a, {"id": a, "name": uname.get(a, "—"), "created": 0, "won": 0, "revenue": 0.0})
        ex[a]["created"] += 1
    for q in won_in:
        a = q.get("assigned_to") or q.get("created_by")
        if not a:
            continue
        ex.setdefault(a, {"id": a, "name": uname.get(a, "—"), "created": 0, "won": 0, "revenue": 0.0})
        ex[a]["won"] += 1
        ex[a]["revenue"] += _value(q)
    executives = sorted(
        [{**v, "revenue": round(v["revenue"], 2)} for v in ex.values()],
        key=lambda x: (-x["revenue"], -x["won"], -x["created"]))

    # Clients
    cl = {}
    for q in created_in:
        name = (q.get("client_snapshot") or {}).get("name") or "—"
        key = q.get("client_id") or name
        cl.setdefault(key, {"name": name, "count": 0, "revenue": 0.0})
        cl[key]["count"] += 1
    for q in won_in:
        name = (q.get("client_snapshot") or {}).get("name") or "—"
        key = q.get("client_id") or name
        cl.setdefault(key, {"name": name, "count": 0, "revenue": 0.0})
        cl[key]["revenue"] += _value(q)
    clients = sorted(
        [{**v, "revenue": round(v["revenue"], 2)} for v in cl.values()],
        key=lambda x: (-x["revenue"], -x["count"]))[:10]

    # Packages most sold (won)
    pk = {}
    for q in won_in:
        if q.get("type") == "servicios":
            continue
        name = (q.get("package_snapshot") or {}).get("name") or q.get("custom_title") or "Programa personalizado"
        pk.setdefault(name, {"name": name, "count": 0, "revenue": 0.0})
        pk[name]["count"] += 1
        pk[name]["revenue"] += _value(q)
    packages = sorted(
        [{**v, "revenue": round(v["revenue"], 2)} for v in pk.values()],
        key=lambda x: (-x["count"], -x["revenue"]))[:10]

    # Services most sold (line items in won quotations)
    sv = {}
    for q in won_in:
        for it in (q.get("items") or []):
            cat = it.get("category")
            if it.get("kind") == "service" or cat in _SERVICE_CATS:
                name = it.get("name") or it.get("label") or "Servicio"
                sv.setdefault(name, {"name": name, "count": 0, "revenue": 0.0})
                sv[name]["count"] += int(it.get("qty", 1) or 1)
                sv[name]["revenue"] += float(it.get("subtotal", 0) or 0)
    services = sorted(
        [{**v, "revenue": round(v["revenue"], 2)} for v in sv.values()],
        key=lambda x: (-x["count"], -x["revenue"]))[:10]

    # Lost quotations
    lost = sorted(lost_in, key=lambda q: _won_date(q), reverse=True)[:50]
    lost_list = [{
        "code": q.get("code"), "client": (q.get("client_snapshot") or {}).get("name", "—"),
        "amount": _value(q), "reason": q.get("lost_reason") or "",
        "date": _won_date(q)[:10],
    } for q in lost]

    return {
        "period": period, "days": days, "currency": currency,
        "revenue_total": revenue_total, "collected_total": collected_total,
        "trend": trend,
        "conversion": {"total": total_created, "won": won_created, "lost": lost_created, "rate": conv_rate},
        "executives": executives, "clients": clients,
        "packages": packages, "services": services,
        "lost": lost_list,
    }


@router.get("/stats/sales")
async def sales_stats(period: str = "month", user: dict = Depends(require_roles("company_admin"))):
    db = get_db()
    return await _compute(db, user["tenant_id"], period)


def _money(v, ccy):
    return f"${float(v or 0):,.2f} {ccy}"


@router.get("/stats/sales/export")
async def export_sales_stats(period: str = "month", user: dict = Depends(require_roles("company_admin"))):
    db = get_db()
    data = await _compute(db, user["tenant_id"], period)
    ccy = data["currency"]
    wb = openpyxl.Workbook()
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="185FA5")

    def _sheet(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font = bold; c.fill = fill
        for r in rows:
            ws.append(r)
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 24
        return ws

    # Resumen
    ws = wb.active; ws.title = "Resumen"
    period_label = {"week": "Última semana", "month": "Último mes", "quarter": "Último trimestre", "year": "Último año"}.get(period, period)
    summary = [
        ["Routiq — Reporte de ventas", ""],
        ["Período", period_label],
        ["Generado", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["", ""],
        ["Ingresos (ganadas)", _money(data["revenue_total"], ccy)],
        ["Cobrado", _money(data["collected_total"], ccy)],
        ["Cotizaciones creadas", data["conversion"]["total"]],
        ["Ganadas", data["conversion"]["won"]],
        ["Perdidas", data["conversion"]["lost"]],
        ["Tasa de conversión", f"{data['conversion']['rate']}%"],
    ]
    for row in summary:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 28

    _sheet("Ejecutivos", ["Ejecutivo", "Creadas", "Cerradas", "Monto vendido"],
           [[e["name"], e["created"], e["won"], _money(e["revenue"], ccy)] for e in data["executives"]])
    _sheet("Clientes", ["Cliente", "Cotizaciones", "Compra total"],
           [[c["name"], c["count"], _money(c["revenue"], ccy)] for c in data["clients"]])
    _sheet("Paquetes", ["Paquete", "Vendidos", "Ingresos"],
           [[p["name"], p["count"], _money(p["revenue"], ccy)] for p in data["packages"]])
    _sheet("Servicios", ["Servicio", "Vendidos", "Ingresos"],
           [[s["name"], s["count"], _money(s["revenue"], ccy)] for s in data["services"]])
    _sheet("Perdidas", ["Folio", "Cliente", "Monto", "Motivo", "Fecha"],
           [[l["code"], l["client"], _money(l["amount"], ccy), l["reason"] or "—", l["date"]] for l in data["lost"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    return StreamingResponse(
        buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="routiq-ventas-{period}-{stamp}.xlsx"'},
    )
