"""Bulk catalog import via Excel (.xlsx).

Provides a downloadable template (Paquetes / Tours / Traslados) and an import
endpoint that validates row-by-row and returns a per-row report of what was
imported and what failed. Admin-only.
"""
import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse

from database import get_db, new_id, now_iso
from auth import require_roles

log = logging.getLogger("routiq.catalog_import")
router = APIRouter()

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
UNITS = {"per_person", "per_group", "per_day", "per_access"}

# Itinerario día a día: dia_1_titulo, dia_1_descripcion ... dia_10_titulo, dia_10_descripcion
ITIN_COLS = [c for i in range(1, 11) for c in (f"dia_{i}_titulo", f"dia_{i}_descripcion")]
PKG_BASE_COLS = ["code", "name", "nights", "description", "image_url", "includes", "excludes",
                 "hotel_name", "sencilla", "doble", "triple", "cuadruple", "menor"]
PKG_COLS = PKG_BASE_COLS + ITIN_COLS
SVC_COLS = ["name", "description", "net_price", "public_price", "unit", "image_url"]

# Hojas de servicios -> categoría, y prefijo de contadores en el reporte de import.
SVC_SHEETS = (("Tours", "tour"), ("Traslados", "traslado"), ("Accesos", "acceso"), ("Extras", "extra"))
CAT_KEY = {"tour": "tours", "traslado": "traslados", "acceso": "accesos", "extra": "extras"}

# Multi-hotel example: PKG-001 has THREE hotels. The first row carries all the
# package-level data + image_url + first hotel + itinerario; the next rows repeat the
# code and only fill the hotel columns (package columns left blank).
PKG_EXAMPLES = [
    ["PKG-001", "Tequila Express 2N", 2, "Tour a Tequila con cata",
     "https://misitio.com/img/tequila.jpg",
     "Hospedaje;Desayunos;Tour", "Vuelos;Propinas",
     "Hotel Solar de las Ánimas", 3500, 2400, 2100, 1900, 1500,
     "Día 1: Llegada a Guadalajara", "Traslado al hotel y tarde libre en el centro histórico (check-in 15:00).",
     "Día 2: Ruta del Tequila", "Salida 09:00 al pueblo mágico de Tequila, visita a destilería con cata y comida."],
    ["PKG-001", "", "", "", "", "", "",
     "Hotel Matices de Amatitán", 3900, 2700, 2300, 2050, 1600],
    ["PKG-001", "", "", "", "", "", "",
     "Hotel Casa Salles", 3200, 2200, 1950, 1750, 1450],
]
TOUR_EXAMPLES = [["City Tour Guadalajara", "Recorrido por el centro histórico", 600, 0, "per_person", ""]]
TRASLADO_EXAMPLES = [["Traslado Aeropuerto-Hotel", "Sedán privado hasta 4 pax", 800, 0, "per_group", ""]]
ACCESO_EXAMPLES = [["Acceso Parque Temático", "Entrada general de día completo", 450, 0, "per_access", ""]]
EXTRA_EXAMPLES = [["Seguro de viaje", "Cobertura médica básica", 120, 0, "per_person", ""]]


def _coerce_num(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        raise ValueError("valor numérico inválido")


def _split_list(v):
    if not v:
        return []
    return [p.strip() for p in str(v).replace("\n", ";").replace(",", ";").split(";") if p.strip()]


# ---------------------------------------------------------------------------
# Template download
# ---------------------------------------------------------------------------
@router.get("/catalog/template")
async def download_template(user: dict = Depends(require_roles("company_admin"))):
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="185FA5")

    def build(ws, cols, examples):
        ws.append(cols)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        for ex in examples:
            ws.append(ex)
        for i, _ in enumerate(cols, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    ws1 = wb.active
    ws1.title = "Paquetes"
    build(ws1, PKG_COLS, PKG_EXAMPLES)
    build(wb.create_sheet("Tours"), SVC_COLS, TOUR_EXAMPLES)
    build(wb.create_sheet("Traslados"), SVC_COLS, TRASLADO_EXAMPLES)
    build(wb.create_sheet("Accesos"), SVC_COLS, ACCESO_EXAMPLES)
    build(wb.create_sheet("Extras"), SVC_COLS, EXTRA_EXAMPLES)

    # Instructions sheet
    info = wb.create_sheet("Instrucciones")
    info.append(["Cómo usar esta plantilla"])
    info["A1"].font = Font(bold=True, size=14)
    for line in [
        "",
        "1. Llena cada hoja: Paquetes, Tours y Traslados.",
        "2. NO borres la fila de encabezados (fila 1). Las filas de ejemplo puedes reemplazarlas o borrarlas.",
        "",
        "PAQUETES — columnas obligatorias: code, name, nights.",
        "   • description: texto de presentación del paquete.",
        "   • image_url: URL de la imagen de portada del paquete (https://...). Opcional; déjala vacía si no tienes.",
        "   • includes/excludes: separa cada elemento con punto y coma (;).",
        "   • sencilla/doble/triple/cuadruple/menor: precios por ocupación de ESE hotel (numéricos).",
        "     Deja el precio en 0 si esa ocupación NO está disponible en el hotel (no se mostrará al cotizar).",
        "",
        "ITINERARIO DÍA A DÍA (opcional, solo en la PRIMERA fila del paquete):",
        "   • Columnas dia_1_titulo, dia_1_descripcion ... hasta dia_10_titulo, dia_10_descripcion.",
        "   • Ejemplo: dia_1_titulo='Día 1: Llegada' | dia_1_descripcion='Traslado al hotel, tarde libre.'",
        "   • Puedes incluir horarios en la descripción (ej. 'Salida 09:00...'). Se cargan automáticamente al paquete.",
        "",
        "MÚLTIPLES HOTELES EN UN MISMO PAQUETE  ← (formato correcto):",
        "   • Usa UNA FILA POR HOTEL repitiendo el MISMO 'code' en cada fila.",
        "   • En la PRIMERA fila del paquete escribe todos los datos (code, name, nights,",
        "     description, image_url, includes, excludes) y los datos del primer hotel.",
        "   • En las filas siguientes repite SOLO el 'code' y llena las columnas del hotel",
        "     (hotel_name, sencilla, doble, triple, cuadruple, menor). Deja vacías name, nights,",
        "     description, image_url, includes y excludes: se toman de la primera fila.",
        "   • También funciona si dejas la columna 'code' vacía en las filas de hoteles adicionales:",
        "     se asignarán automáticamente al paquete de la fila anterior.",
        "",
        "   Ejemplo (hoja 'Paquetes', filas 2 a 4) — el paquete PKG-001 con 3 hoteles:",
        "      Fila 2:  code=PKG-001 | name=Tequila Express 2N | nights=2 | ... | hotel_name=Hotel Solar de las Ánimas | precios...",
        "      Fila 3:  code=PKG-001 | (resto vacío)                          | hotel_name=Hotel Matices de Amatitán | precios...",
        "      Fila 4:  code=PKG-001 | (resto vacío)                          | hotel_name=Hotel Casa Salles         | precios...",
        "",
        "   Resultado en el sistema (Catálogo → Paquetes):",
        "      • Se crea UN solo paquete 'PKG-001 — Tequila Express 2N'.",
        "      • En su tarjeta verás la etiqueta '3 hoteles disponibles'.",
        "      • Al cotizar, el ejecutivo elige entre los 3 hoteles y cada uno usa sus propios precios por ocupación.",
        "",
        "ACTUALIZAR PRECIOS / TARIFAS DE TEMPORADA (sin borrar):",
        "   • Si el 'code' YA EXISTE en tu catálogo, el paquete se ACTUALIZA (no se duplica):",
        "     se sobrescriben nombre, noches, descripción, imagen, incluye/no incluye y TODOS los hoteles",
        "     con lo que traiga el Excel. Ideal para actualizar tarifas cada año o por temporada.",
        "   • Si el 'code' es nuevo, se crea un paquete nuevo.",
        "",
        "TOURS / TRASLADOS / ACCESOS / EXTRAS — columna obligatoria: name.",
        "   • Cada hoja corresponde a una categoría de servicio (Tours, Traslados, Accesos, Extras).",
        "   • description: descripción del servicio.",
        "   • net_price = costo; public_price = precio de venta (déjalo en 0 para autocalcular con tu margen).",
        "   • unit: per_person, per_group, per_day o per_access.",
        "   • image_url: URL de imagen del servicio (https://...). Opcional.",
        "   • ACTUALIZAR sin duplicar: si coincide el NOMBRE (tours/accesos/extras) o el NOMBRE + UNIDAD (traslados),",
        "     el servicio se actualiza en lugar de crear uno nuevo.",
        "   • IMPORTACIÓN POR CATEGORÍA INDEPENDIENTE: puedes subir un Excel con SOLO una hoja llena",
        "     (p.ej. solo Tours) y las demás categorías NO se verán afectadas.",
        "",
        "3. Guarda el archivo y súbelo desde el panel: Catálogo → Importar Excel (o Servicios → Importar Excel).",
        "   Al terminar verás un resumen: 'X paquetes nuevos, Y actualizados, Z hoteles · tours nuevos/actualizados · traslados nuevos/actualizados'.",
    ]:
        info.append([line])
    info.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="routiq-catalogo-template.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Export current catalog (same format as the template)
# ---------------------------------------------------------------------------
@router.get("/catalog/export")
async def export_catalog(user: dict = Depends(require_roles("company_admin"))):
    db = get_db()
    tenant_id = user["tenant_id"]
    packages = await db.packages.find({"tenant_id": tenant_id}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    services = await db.services.find({"tenant_id": tenant_id}, {"_id": 0}).sort("created_at", -1).to_list(2000)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="185FA5")

    def header(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
        for i, _ in enumerate(cols, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    ws_pkg = wb.active
    ws_pkg.title = "Paquetes"
    header(ws_pkg, PKG_COLS)
    for p in packages:
        hotels = p.get("hotels") or [{}]
        itin = {int(d.get("day", 0)): d for d in (p.get("itinerary") or []) if d}
        itin_cells = []
        for i in range(1, 11):
            dd = itin.get(i, {})
            itin_cells += [dd.get("title", ""), dd.get("description", "")]
        for i, hotel in enumerate(hotels):
            occ = hotel.get("prices_by_occupancy", {}) if hotel else {}
            if i == 0:
                ws_pkg.append([
                    p.get("code", ""), p.get("name", ""), p.get("nights", 0), p.get("description", ""),
                    p.get("image_url", ""),
                    ";".join(p.get("includes", [])), ";".join(p.get("excludes", [])),
                    hotel.get("name", "") if hotel else "",
                    occ.get("sencilla", 0), occ.get("doble", 0), occ.get("triple", 0),
                    occ.get("cuadruple", 0), hotel.get("minor_price", 0) if hotel else 0,
                    *itin_cells,
                ])
            else:
                # Continuation row: repeat the code, leave package columns blank, add the extra hotel.
                ws_pkg.append([
                    p.get("code", ""), "", "", "", "", "", "",
                    hotel.get("name", ""),
                    occ.get("sencilla", 0), occ.get("doble", 0), occ.get("triple", 0),
                    occ.get("cuadruple", 0), hotel.get("minor_price", 0),
                ])

    def svc_sheet(title, category):
        ws = wb.create_sheet(title)
        header(ws, SVC_COLS)
        for s in services:
            if s.get("category") != category:
                continue
            ws.append([s.get("name", ""), s.get("description", ""),
                       s.get("net_price", 0), s.get("public_price", 0),
                       s.get("unit", "per_group"), s.get("image_url", "")])

    for title, cat in SVC_SHEETS:
        svc_sheet(title, cat)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime, timezone
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    return StreamingResponse(
        buf, media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="routiq-catalogo-{stamp}.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
@router.post("/catalog/import")
async def import_catalog(file: UploadFile = File(...), user: dict = Depends(require_roles("company_admin"))):
    name = (file.filename or "").lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sube un archivo .xlsx (usa la plantilla descargable).")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande (máx 5 MB).")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el Excel. ¿Está corrupto o no es .xlsx?")

    db = get_db()
    tenant_id = user["tenant_id"]
    company = await db.companies.find_one({"id": tenant_id}, {"_id": 0, "pricing_config": 1})
    divisor = float((company or {}).get("pricing_config", {}).get("margin_divisor") or 0.76) or 0.76

    errors: list[dict] = []
    imported = {
        "paquetes_nuevos": 0, "paquetes_actualizados": 0, "hoteles": 0,
        "tours_nuevos": 0, "tours_actualizados": 0,
        "traslados_nuevos": 0, "traslados_actualizados": 0,
        "accesos_nuevos": 0, "accesos_actualizados": 0,
        "extras_nuevos": 0, "extras_actualizados": 0,
    }

    # --- Paquetes (grouped by code; one row per hotel) ---
    if "Paquetes" in wb.sheetnames:
        ws = wb["Paquetes"]
        # 1) Group rows into packages. Rows sharing a 'code' (or with a blank
        #    'code' that continues the previous package) become one package with
        #    multiple hotels. Package-level fields come from the first row.
        groups: dict[str, dict] = {}
        order: list[str] = []
        current_code: str | None = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            d = dict(zip(PKG_COLS, list(row) + [None] * len(PKG_COLS)))
            code = str(d.get("code") or "").strip()
            if not code:
                # Continuation row -> belongs to the previous package.
                if not current_code:
                    errors.append({"sheet": "Paquetes", "row": idx,
                                   "message": "falta 'code' y no hay un paquete previo al que pertenezca esta fila"})
                    continue
                code = current_code
            else:
                current_code = code
            grp = groups.get(code)
            if grp is None:
                grp = {"code": code, "base": d, "first_row": idx, "hotel_rows": []}
                groups[code] = grp
                order.append(code)
            # Collect the hotel defined on this row (if any).
            if str(d.get("hotel_name") or "").strip():
                grp["hotel_rows"].append((idx, d))

        # 2) Validate and insert one package per group.
        for code in order:
            grp = groups[code]
            d = grp["base"]
            first_row = grp["first_row"]
            try:
                pname = str(d.get("name") or "").strip()
                if not pname:
                    raise ValueError("name es obligatorio (en la primera fila del paquete)")
                nights = int(_coerce_num(d.get("nights"), 0))
                if nights <= 0:
                    raise ValueError("nights debe ser un entero > 0 (en la primera fila del paquete)")
                existing = await db.packages.find_one({"tenant_id": tenant_id, "code": code}, {"_id": 0, "id": 1})
                hotels = []
                for (hidx, hd) in grp["hotel_rows"]:
                    try:
                        hotels.append({
                            "name": str(hd["hotel_name"]).strip(), "category": "",
                            "prices_by_occupancy": {
                                "sencilla": _coerce_num(hd["sencilla"]), "doble": _coerce_num(hd["doble"]),
                                "triple": _coerce_num(hd["triple"]), "cuadruple": _coerce_num(hd["cuadruple"]),
                            },
                            "minor_price": _coerce_num(hd["menor"]), "season_prices": {},
                        })
                    except Exception as he:
                        errors.append({"sheet": "Paquetes", "row": hidx,
                                       "message": f"hotel '{hd.get('hotel_name')}': {he}"})
                fields = {
                    "name": pname, "nights": nights,
                    "description": str(d.get("description") or ""),
                    "image_url": str(d.get("image_url") or "").strip(),
                    "hotels": hotels,
                    "includes": _split_list(d.get("includes")), "excludes": _split_list(d.get("excludes")),
                }
                # Itinerario día a día (dia_1_titulo/dia_1_descripcion ... dia_10_*) — solo desde la fila base.
                itinerary = []
                for i in range(1, 11):
                    t = str(d.get(f"dia_{i}_titulo") or "").strip()
                    desc = str(d.get(f"dia_{i}_descripcion") or "").strip()
                    if t or desc:
                        itinerary.append({"day": i, "title": t, "description": desc})
                if itinerary:
                    fields["itinerary"] = itinerary
                if existing:
                    # Update existing package (overwrite data + hotels) — keeps id/created_at/status.
                    await db.packages.update_one(
                        {"tenant_id": tenant_id, "code": code},
                        {"$set": {**fields, "updated_at": now_iso()}})
                    imported["paquetes_actualizados"] += 1
                else:
                    doc = {
                        "id": new_id(), "tenant_id": tenant_id, "created_at": now_iso(),
                        "code": code, "itinerary": [], "seasons": [],
                        "allowed_start_days": [], "special_departure_dates": [], "status": "active",
                        **fields,
                    }
                    await db.packages.insert_one(dict(doc))
                    imported["paquetes_nuevos"] += 1
                imported["hoteles"] += len(hotels)
            except Exception as e:
                errors.append({"sheet": "Paquetes", "row": first_row, "message": str(e)})

    # --- Servicios por categoría (Tours / Traslados / Accesos / Extras) ---
    # Upsert por clave de unicidad: por defecto (name, category); traslados -> (name, unit, category).
    for sheet, category in SVC_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        key = CAT_KEY[category]
        ws = wb[sheet]
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            d = dict(zip(SVC_COLS, list(row) + [None] * len(SVC_COLS)))
            try:
                sname = str(d["name"] or "").strip()
                if not sname:
                    raise ValueError("name es obligatorio")
                net = _coerce_num(d["net_price"])
                pub = _coerce_num(d["public_price"])
                if pub <= 0:
                    pub = round(net / divisor, 2) if net else 0.0
                unit = str(d.get("unit") or "per_group").strip()
                if unit not in UNITS:
                    unit = "per_group"
                fields = {
                    "description": str(d.get("description") or ""),
                    "net_price": net, "public_price": pub, "unit": unit,
                    "per_person": unit == "per_person",
                    "image_url": str(d.get("image_url") or "").strip(),
                }
                match = {"tenant_id": tenant_id, "category": category, "name": sname}
                if category == "traslado":
                    match["unit"] = unit
                existing = await db.services.find_one(match, {"_id": 0, "id": 1})
                if existing:
                    await db.services.update_one(match, {"$set": {**fields, "updated_at": now_iso()}})
                    imported[f"{key}_actualizados"] += 1
                else:
                    doc = {
                        "id": new_id(), "tenant_id": tenant_id, "created_at": now_iso(),
                        "name": sname, "category": category, "status": "active", **fields,
                    }
                    await db.services.insert_one(dict(doc))
                    imported[f"{key}_nuevos"] += 1
            except Exception as e:
                errors.append({"sheet": sheet, "row": idx, "message": str(e)})

    wb.close()
    total = (imported["paquetes_nuevos"] + imported["paquetes_actualizados"]
             + sum(imported[f"{k}_nuevos"] + imported[f"{k}_actualizados"]
                   for k in ("tours", "traslados", "accesos", "extras")))
    return {"ok": True, "imported": imported, "total_imported": total, "errors": errors, "error_count": len(errors)}
